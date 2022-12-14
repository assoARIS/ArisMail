# -*- coding: utf-8 -*-
"""
Created on Tue Feb 22 20:55:40 2022

@author: JMichel
"""

from openpyxl import load_workbook
from datetime import datetime

''' fichiers Excel '''
fchExcelR = "FICHIER_ADHERENTS_au_30_NOVEMBRE_2022.xlsx"
feuilExcelR = "FICHIER"
fchExcelW = "ARIS-MailingList-20221202.xlsx"
feuilExcelW = "Adherents"
#

''' variables '''
global workBookR , workSheetR
global workBookW , workSheetW

''' fonctions '''
#
def clear():
    print("\033[H\033[J")
#
def setGlobal():
    ''' variable de controle d'execution '''
    global chk
    chk = True
#
def openExcelFile():
    ''' ouverture du fichier Excel et selection du feuillet des adherents '''
    global chk
    if( chk == True ):
        try:
            global workBookR , workSheetR
            workBookR = load_workbook(fchExcelR)
            workSheetR = workBookR[feuilExcelR]
        except:
            print("! erreur de connexion au fichier Excel Adherents !")
            chk = False
    #
    if( chk == True ):
        try:
            global workBookW , workSheetW
            workBookW = load_workbook(fchExcelW)
            workSheetW = workBookW[feuilExcelW]
        except:
            print("! erreur de connexion au fichier Excel MailingList !")
            chk = False
#
def copyData():
    ''' processus de recopie des donnees mails des adherents '''
    if( chk == True ):
        try:
            i = 11
            j = 2
            while(workSheetR["L"+str(i)].value != None):
                ''' recuperation des nom prenom de l'adherent '''
                nom = workSheetR["L"+str(i)].value
                if(workSheetR["M"+str(i)].value != None):
                    prenom = workSheetR["M"+str(i)].value
                else:
                    prenom = ""
                ''' recuperation de l'adresse mail de l'adherent '''
                if(workSheetR["AB"+str(i)].value != None
                and workSheetR["AB"+str(i)].value.find("@") != -1):
                    adresseMail = workSheetR["AB"+str(i)].value
                else:
                    adresseMail = "??"
                ''' recuperation du statut de l'adherent '''
                if(workSheetR["A"+str(i)].value != None):
                    statut = workSheetR["A"+str(i)].value
                else:
                    statut = "??"
                if(workSheetR["R"+str(i)].value != None
                and workSheetR["O"+str(i)].value != None):
                    dateNaissance =  str(workSheetR["O"+str(i)].value).strip() + "/"
                    dateNaissance += str(workSheetR["P"+str(i)].value).strip() + "/" 
                    dateNaissance += str(workSheetR["Q"+str(i)].value).strip()
                else:
                    dateNaissance = "00/00/0000"
                ''' recopie des donnees adherent '''
                workSheetW["A"+str(j)].value = nom
                workSheetW["B"+str(j)].value = prenom
                workSheetW["C"+str(j)].value = adresseMail
                workSheetW["D"+str(j)].value = statut
                workSheetW["E"+str(j)].value = getNaissance(dateNaissance)
                ''' affichage des donnees adherent '''
                adh = "i:"+str(i)+" / j:"+str(j)+" - "+nom +" "+prenom+" ("+statut+")  : "+adresseMail
                adh += "  -- "+dateNaissance
                adh += "  -- "+getNaissance(dateNaissance)
                print(adh)
                i = i + 1
                j = j + 1
        except:
            adh = "i:"+str(i)+" / j:"+str(j)+" - "+nom +" "+prenom+" ("+statut+")  : "+adresseMail
            adh += "  -- "+dateNaissance
            adh += "  -- "+getNaissance(dateNaissance)
            print("! erreur sur donnees adherent : " + adh + " !")
#
def closeExcelFile(): 
    ''' enregistrement & fermeture des fichiers Excel '''
    try:
        workBookR.close()
        workBookW.save(fchExcelW)
        workBookW.close()
    except:
        print("! erreur d'enregistrement / fermeture fichier Excel !")
#
def getNaissance(dateNaissance):
    if(dateNaissance == "//"):
        dateNaissance = "00/00/0000"
    lstNaissance = dateNaissance.split("/")
    if(len(lstNaissance[0]) == 1):
       lstNaissance[0] = "0" + lstNaissance[0]
    if(len(lstNaissance[1]) == 1):
       lstNaissance[1] = "0" + lstNaissance[1]
    dateNaissance = lstNaissance[0] + "/" + lstNaissance[1] + "/" + lstNaissance[2]
    return(dateNaissance)
#
        
    
''' execution '''
#
def main():
    clear()
    setGlobal()
    print("- start process -\n")
    if(chk == True):
        openExcelFile()
    if(chk == True):
        copyData()
    closeExcelFile()
    print("\n- end process -")
#
if(__name__ == "__main__"):
  main()
#
