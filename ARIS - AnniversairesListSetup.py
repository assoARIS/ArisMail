#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Thu Dec  8 20:43:29 2022

@author: JMichel
"""

from openpyxl import load_workbook
from datetime import datetime
from json import loads, dumps

''' fichiers Excel '''
fchExcelR = "./ARIS-MailingList-20221202.xlsx"
feuilExcelR = "Adherents"
fchExcelW = ""
feuilExcelW = ""
#

''' fichier Texte '''
fchText = "./ARIS-AnniversairesListe.txt"
#

''' variables '''
#global workBookR , workSheetR
#global workBookW , workSheetW
#

''' fonctions '''
def clear():
    print("\033[H\033[J")
#
def setGlobal():
    ''' variable de controle d'execution '''
    global chk
    chk = True
#
def openExcelFile():
    ''' ouverture du fichier Excel et selection du feuillet '''
    global chk
    if( chk == True ):
        try:
            global workBookR , workSheetR
            workBookR = load_workbook(fchExcelR)
            workSheetR = workBookR[feuilExcelR]
        except:
            print("! erreur de connexion au fichier Excel !")
            chk = False
    #
    """
    if( chk == True ):
        try:
            global workBookW , workSheetW
            workBookW = load_workbook(fchExcelW)
            workSheetW = workBookW[feuilExcelW]
        except:
            print("! erreur de connexion au fichier Excel !")
            chk = False
    """
#
def closeExcelFile(): 
    ''' enregistrement & fermeture des fichiers Excel '''
    try:
        workBookR.close()
        """
        workBookW.save(fchExcelW)
        workBookW.close()
        """
    except:
        print("! erreur d'enregistrement / fermeture fichier Excel !")
#
def copyData():
    ''' processus de recopie des donnees mails des adherents '''
    global dicoAdherent, dicoNaissances
    dicoNaissances = {}
    dicoAdherent = {}
    if( chk == True ):
        try:
            i = 2
            j = 1
            while(workSheetR["A"+str(i)].value != None):
                ''' recuperation des nom prenom de l'adherent '''
                nom = workSheetR["A"+str(i)].value
                if(workSheetR["B"+str(i)].value != None):
                    prenom = workSheetR["B"+str(i)].value
                else:
                    prenom = ""
                ''' recuperation de l'adresse mail de l'adherent '''
                if(workSheetR["C"+str(i)].value != None
                and workSheetR["C"+str(i)].value.find("@") != -1):
                    adresseMail = workSheetR["C"+str(i)].value
                ''' recuperation du statut de l'adherent '''
                if(workSheetR["D"+str(i)].value != None):
                    statut = workSheetR["D"+str(i)].value
                ''' recuperation de la date de naissance de l'adherent '''
                if(workSheetR["E"+str(i)].value != None
                and workSheetR["E"+str(i)].value != "00/00/000"):
                    dateNaissance = workSheetR["E"+str(i)].value
                #
                ''' recopie des donnees adherent dans le dictionnaire '''
                iStr = getIntStr(j)
                listAdherent = [(iStr,(nom,prenom,adresseMail,statut,dateNaissance))]
                dicoAdherent = dict(listAdherent)
                dicoNaissances.update(dicoAdherent)
                i = i + 1
                j = j + 1
                jsonNaissances = dumps(dicoNaissances)
                writeText("")
                appendText(jsonNaissances)
        except:
            adh = "j:"+str(j)+" - "+nom +" "+prenom+" ("+statut+")  : "+adresseMail
            adh += "  -- "+dateNaissance
            print("! erreur sur donnees adherent : " + adh + " !")
#
def writeText(txt):
    fchW = open(fchText,"w")
    fchW.write(txt)
    fchW.close()
#
def appendText(txt):
    fchA = open(fchText,"a")
    fchA.write(txt)
    fchA.close()
#
def readText():
    fchR = open(fchText,"r")
    txt = fchR.read()
    fchR.close()
    return(txt)
#
def getIntStr(i):
    iStr = str(i)
    if(len(iStr) == 1):
        iStr = "00"+iStr
    elif(len(iStr) == 2):
        iStr = "0"+iStr
    return(iStr)
    

''' execution '''
#
def main():
    clear()
    setGlobal()
    print("- start process -\n")
    if(chk == True):
        openExcelFile()
    if(chk == True):
        copyData()
    closeExcelFile()
    print("\n- end process -")
#
if(__name__ == "__main__"):
  main()
#